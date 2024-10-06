import pandas as pd 
import paramiko
from io import StringIO
import socket
import openpyxl 



def connect_to_server(hostname, port, username, password, commands):
    try:
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        client.connect(hostname=hostname, port=port, username=username, password=password, timeout=10)
        
        outputs = []
        for c in commands:
            stdin, stdout, stderr = client.exec_command(c)
            output = stdout.read().decode()
            outputs.append(output.strip())
        client.close()
        return outputs

    except paramiko.AuthenticationException:
        return "Wrong password"
    except (paramiko.SSHException, socket.timeout) as e:
        return f"Connection error: {str(e)}"
    except Exception as e:
        return f"{str(e)}"

def parse_df_output(output):
    try:
        df = pd.read_csv(StringIO(output), sep='\s+')
        return df
    except pd.errors.EmptyDataError:
        return None
    except Exception as e:
        print(f"Error parsing output: {e}")
        return None



def process_servers(servers_df, template_path, output_path):
    messages = [] 
    commands = [
        "df -h --total", 
        "mpstat 1 1 | awk 'NR==4{print $3 + $4 + $5 + $6}'",
        "free -h | awk 'NR==2{print ($3/$2)*100}'"
    ]

    try:
        successful_connections = False 
        wb = openpyxl.load_workbook(template_path)
        ws = wb['data']

        for index, row in servers_df.iterrows():
            hostname = row['servers']  
            port = row['port']
            username = row['username']
            password = row['password']
            
            result = connect_to_server(hostname, port, username, password, commands)
            
            if isinstance(result, str):  
                messages.append(f"Failed to connect to {hostname, username}: {result}")
                continue  
            
            df = parse_df_output(result[0])
            
            if df is not None and not df.empty:
                successful_connections = True
                try:
                    disk_usage = df.iloc[-1]["Use%"] if "Use%" in df.columns else 'N/A'
                    cpu = float(result[1]) / 100  
                    ram_usage = float(result[2].strip()) / 100  

                    
                    for excel_row in range(2, ws.max_row + 1):
                        server_name = ws.cell(row=excel_row, column=1).value
                        if server_name == hostname:
                            
                            ws.cell(row=excel_row, column=3).value = float(disk_usage.strip('%')) / 100  
                            ws.cell(row=excel_row, column=5).value = cpu                                 
                            ws.cell(row=excel_row, column=7).value = ram_usage                           
                           

                            ws.cell(row=excel_row, column=3).number_format = '0.00%'  
                            ws.cell(row=excel_row, column=5).number_format = '0.00%'  
                            ws.cell(row=excel_row, column=7).number_format = '0.00%'  
                except Exception as e:
                    messages.append(f"Error processing data from {hostname}: {e}")
            else:
                messages.append(f"Invalid output from {hostname}, skipping.")
        
        if not successful_connections:
            messages.append("Failed to connect to all servers. No file generated.")
            return messages
        

        wb.save(template_path) 
        wb.save(output_path)  
       
        return messages

    except Exception as e:
        return [f"Error updating Excel template: {str(e)}"]
